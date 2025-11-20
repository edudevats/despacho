"""
Rutas de Movimientos
Registro y gestión de ingresos y egresos con jerarquía de 4 niveles
"""
from flask import render_template, redirect, url_for, flash, request, current_app, send_file
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from datetime import datetime, date
from sqlalchemy import or_, and_, extract
from decimal import Decimal
import os
from io import BytesIO

from app.movimientos import bp
from app.auth.decoradores import misma_sucursal_required, login_required_with_role
from app.auth.auditoria import registrar_auditoria
from app.models import (Ingreso, Egreso, CategoriaIngreso, CategoriaEgreso, Sucursal, Organizacion, Usuario, db)
from app.movimientos.forms import IngresoForm, EgresoForm


@bp.route('/registro', methods=['GET', 'POST'])
@login_required
@misma_sucursal_required
def registro():
    """
    Pantalla unificada de registro de movimientos con tabs
    Solo usuarios de sucursales pueden registrar movimientos
    """
    ingreso_form = IngresoForm()
    egreso_form = EgresoForm()

    # Obtener sucursal del usuario actual
    sucursal_id = current_user.sucursal_id

    if not sucursal_id:
        flash('Solo los usuarios asignados a una sucursal pueden registrar movimientos.', 'warning')
        return redirect(url_for('main.index'))

    # Cargar categorías disponibles para esta sucursal
    # Incluye categorías de organización (compartidas) + categorías específicas de la sucursal
    sucursal = Sucursal.query.get(sucursal_id)

    # Categorías de ingreso: org + sucursal
    categorias_ingreso_org = CategoriaIngreso.query.filter_by(
        organizacion_id=sucursal.organizacion_id,
        activa=True
    ).order_by(CategoriaIngreso.nombre).all()

    categorias_ingreso_suc = CategoriaIngreso.query.filter_by(
        sucursal_id=sucursal_id,
        activa=True
    ).order_by(CategoriaIngreso.nombre).all()

    todas_categorias_ingreso = categorias_ingreso_org + categorias_ingreso_suc
    ingreso_form.categoria_id.choices = [(c.id, c.nombre) for c in todas_categorias_ingreso]

    # Categorías de egreso: org + sucursal
    categorias_egreso_org = CategoriaEgreso.query.filter_by(
        organizacion_id=sucursal.organizacion_id,
        activa=True
    ).order_by(CategoriaEgreso.nombre).all()

    categorias_egreso_suc = CategoriaEgreso.query.filter_by(
        sucursal_id=sucursal_id,
        activa=True
    ).order_by(CategoriaEgreso.nombre).all()

    todas_categorias_egreso = categorias_egreso_org + categorias_egreso_suc
    egreso_form.categoria_id.choices = [(c.id, c.nombre) for c in todas_categorias_egreso]

    # Procesar formulario de ingreso
    if request.method == 'POST' and request.form.get('tipo') == 'ingreso':
        if ingreso_form.validate():
            # Verificar permisos
            if current_user.es_secundario_pos() and not current_user.puede_registrar_ingresos:
                flash('No tienes permiso para registrar ingresos.', 'warning')
                return redirect(url_for('movimientos.registro'))

            ingreso = Ingreso(
                sucursal_id=sucursal_id,
                usuario_id=current_user.id,
                categoria_id=ingreso_form.categoria_id.data,
                fecha=ingreso_form.fecha.data,
                monto=ingreso_form.monto.data,
                metodo_pago=ingreso_form.metodo_pago.data,
                notas=ingreso_form.notas.data
            )

            db.session.add(ingreso)
            db.session.commit()

            registrar_auditoria('create', 'ingreso', ingreso.id, {
                'monto': str(ingreso.monto),
                'categoria': ingreso.categoria.nombre
            })

            flash('Ingreso registrado exitosamente.', 'success')
            return redirect(url_for('movimientos.registro'))

    # Procesar formulario de egreso
    if request.method == 'POST' and request.form.get('tipo') == 'egreso':
        if egreso_form.validate():
            # Verificar permisos
            if current_user.es_secundario_pos() and not current_user.puede_registrar_egresos:
                flash('No tienes permiso para registrar egresos.', 'warning')
                return redirect(url_for('movimientos.registro'))

            egreso = Egreso(
                sucursal_id=sucursal_id,
                usuario_id=current_user.id,
                categoria_id=egreso_form.categoria_id.data,
                fecha=egreso_form.fecha.data,
                monto=egreso_form.monto.data,
                metodo_pago=egreso_form.metodo_pago.data,
                notas=egreso_form.notas.data
            )

            # Procesar comprobante si existe
            if egreso_form.comprobante.data:
                file = egreso_form.comprobante.data
                filename = secure_filename(file.filename)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"{timestamp}_{filename}"

                upload_path = os.path.join(
                    current_app.config['UPLOAD_FOLDER'],
                    'comprobantes',
                    str(sucursal_id)
                )
                os.makedirs(upload_path, exist_ok=True)

                file_path = os.path.join(upload_path, filename)
                file.save(file_path)

                egreso.comprobante_filename = filename
                egreso.comprobante_path = file_path

            db.session.add(egreso)
            db.session.commit()

            registrar_auditoria('create', 'egreso', egreso.id, {
                'monto': str(egreso.monto),
                'categoria': egreso.categoria.nombre
            })

            flash('Egreso registrado exitosamente.', 'success')
            return redirect(url_for('movimientos.registro'))

    return render_template(
        'movimientos/registro.html',
        ingreso_form=ingreso_form,
        egreso_form=egreso_form
    )


@bp.route('/listado')
@login_required
def listado():
    """
    Vista unificada de movimientos con búsqueda avanzada
    - Admin del despacho: ve todas las sucursales
    - Director de organización: ve todas las sucursales de su organización
    - Gerente de sucursal: solo su sucursal
    - POS: solo su sucursal
    """
    # Obtener parámetro de organización seleccionada
    organizacion_id_filtro = request.args.get('organizacion_id', type=int)

    # Obtener sucursales según rol del usuario y organización seleccionada
    if current_user.es_admin_despacho():
        organizaciones = Organizacion.query.filter_by(activa=True).order_by(Organizacion.nombre).all()

        # Filtrar sucursales según organización seleccionada
        if organizacion_id_filtro:
            sucursales = Sucursal.query.filter_by(
                organizacion_id=organizacion_id_filtro,
                activa=True
            ).order_by(Sucursal.nombre).all()
        else:
            sucursales = Sucursal.query.filter_by(activa=True).order_by(Sucursal.nombre).all()

        # Admin ve todas las categorías (filtradas por organización si aplica)
        if organizacion_id_filtro:
            # Categorías de la organización seleccionada
            categorias_ingreso_org = CategoriaIngreso.query.filter_by(
                organizacion_id=organizacion_id_filtro,
                activa=True
            ).order_by(CategoriaIngreso.nombre).all()

            # Categorías de las sucursales de esa organización
            sucursales_ids_org = [s.id for s in sucursales]
            categorias_ingreso_suc = CategoriaIngreso.query.filter(
                CategoriaIngreso.sucursal_id.in_(sucursales_ids_org),
                CategoriaIngreso.activa == True
            ).order_by(CategoriaIngreso.nombre).all() if sucursales_ids_org else []

            todas_las_categorias_ingreso = categorias_ingreso_org + categorias_ingreso_suc

            # Lo mismo para egresos
            categorias_egreso_org = CategoriaEgreso.query.filter_by(
                organizacion_id=organizacion_id_filtro,
                activa=True
            ).order_by(CategoriaEgreso.nombre).all()

            categorias_egreso_suc = CategoriaEgreso.query.filter(
                CategoriaEgreso.sucursal_id.in_(sucursales_ids_org),
                CategoriaEgreso.activa == True
            ).order_by(CategoriaEgreso.nombre).all() if sucursales_ids_org else []

            todas_las_categorias_egreso = categorias_egreso_org + categorias_egreso_suc
        else:
            # Todas las categorías
            todas_las_categorias_ingreso = CategoriaIngreso.query.filter_by(activa=True).order_by(CategoriaIngreso.nombre).all()
            todas_las_categorias_egreso = CategoriaEgreso.query.filter_by(activa=True).order_by(CategoriaEgreso.nombre).all()

    elif current_user.es_principal_organizacion():
        # Director ve todas las sucursales de su organización
        sucursales = Sucursal.query.filter_by(
            organizacion_id=current_user.organizacion_id,
            activa=True
        ).order_by(Sucursal.nombre).all()
        organizaciones = [Organizacion.query.get(current_user.organizacion_id)]

        # Ve categorías de su organización + categorías de sus sucursales
        categorias_ingreso_org = CategoriaIngreso.query.filter_by(
            organizacion_id=current_user.organizacion_id,
            activa=True
        ).order_by(CategoriaIngreso.nombre).all()

        sucursales_ids = [s.id for s in sucursales]
        categorias_ingreso_suc = CategoriaIngreso.query.filter(
            CategoriaIngreso.sucursal_id.in_(sucursales_ids),
            CategoriaIngreso.activa == True
        ).order_by(CategoriaIngreso.nombre).all() if sucursales_ids else []

        todas_las_categorias_ingreso = categorias_ingreso_org + categorias_ingreso_suc

        # Lo mismo para egresos
        categorias_egreso_org = CategoriaEgreso.query.filter_by(
            organizacion_id=current_user.organizacion_id,
            activa=True
        ).order_by(CategoriaEgreso.nombre).all()

        categorias_egreso_suc = CategoriaEgreso.query.filter(
            CategoriaEgreso.sucursal_id.in_(sucursales_ids),
            CategoriaEgreso.activa == True
        ).order_by(CategoriaEgreso.nombre).all() if sucursales_ids else []

        todas_las_categorias_egreso = categorias_egreso_org + categorias_egreso_suc

    else:
        # Gerente o POS solo ven su sucursal
        sucursales = [Sucursal.query.get(current_user.sucursal_id)]
        organizaciones = [Organizacion.query.get(current_user.organizacion_id)]

        # Ve categorías de su organización + su sucursal
        categorias_ingreso_org = CategoriaIngreso.query.filter_by(
            organizacion_id=current_user.organizacion_id,
            activa=True
        ).order_by(CategoriaIngreso.nombre).all()

        categorias_ingreso_suc = CategoriaIngreso.query.filter_by(
            sucursal_id=current_user.sucursal_id,
            activa=True
        ).order_by(CategoriaIngreso.nombre).all()

        todas_las_categorias_ingreso = categorias_ingreso_org + categorias_ingreso_suc

        # Lo mismo para egresos
        categorias_egreso_org = CategoriaEgreso.query.filter_by(
            organizacion_id=current_user.organizacion_id,
            activa=True
        ).order_by(CategoriaEgreso.nombre).all()

        categorias_egreso_suc = CategoriaEgreso.query.filter_by(
            sucursal_id=current_user.sucursal_id,
            activa=True
        ).order_by(CategoriaEgreso.nombre).all()

        todas_las_categorias_egreso = categorias_egreso_org + categorias_egreso_suc

    # Obtener parámetros de búsqueda/filtrado
    organizacion_id = request.args.get('organizacion_id', type=int)
    sucursal_id = request.args.get('sucursal_id', type=int)
    tipo = request.args.get('tipo', '')  # '', 'ingreso', 'egreso'
    categoria_id = request.args.get('categoria_id', type=int)
    fecha_inicio = request.args.get('fecha_inicio', '')
    fecha_fin = request.args.get('fecha_fin', '')
    monto_min = request.args.get('monto_min', type=float)
    monto_max = request.args.get('monto_max', type=float)
    metodo_pago = request.args.get('metodo_pago', '')
    buscar_texto = request.args.get('buscar_texto', '')

    # Aplicar restricciones según rol
    if current_user.es_principal_organizacion():
        organizacion_id = current_user.organizacion_id
    elif not current_user.es_admin_despacho():
        sucursal_id = current_user.sucursal_id

    resultados_ingresos = []
    resultados_egresos = []

    # Construir queries base para ingresos
    if tipo == '' or tipo == 'ingreso':
        ingresos_query = Ingreso.query.join(Sucursal)

        # Aplicar filtros
        if sucursal_id:
            ingresos_query = ingresos_query.filter(Ingreso.sucursal_id == sucursal_id)
        elif organizacion_id:
            # Filtrar por organización (todas las sucursales de esa org)
            sucursales_org = [s.id for s in Sucursal.query.filter_by(organizacion_id=organizacion_id).all()]
            if sucursales_org:
                ingresos_query = ingresos_query.filter(Ingreso.sucursal_id.in_(sucursales_org))
        elif current_user.es_principal_organizacion():
            # Director sin filtro específico: todas sus sucursales
            sucursales_org = [s.id for s in sucursales]
            if sucursales_org:
                ingresos_query = ingresos_query.filter(Ingreso.sucursal_id.in_(sucursales_org))
        elif not current_user.es_admin_despacho():
            # Gerente/POS: solo su sucursal
            ingresos_query = ingresos_query.filter(Ingreso.sucursal_id == current_user.sucursal_id)

        if categoria_id:
            ingresos_query = ingresos_query.filter(Ingreso.categoria_id == categoria_id)
        if fecha_inicio:
            ingresos_query = ingresos_query.filter(Ingreso.fecha >= datetime.strptime(fecha_inicio, '%Y-%m-%d').date())
        if fecha_fin:
            ingresos_query = ingresos_query.filter(Ingreso.fecha <= datetime.strptime(fecha_fin, '%Y-%m-%d').date())
        if monto_min:
            ingresos_query = ingresos_query.filter(Ingreso.monto >= monto_min)
        if monto_max:
            ingresos_query = ingresos_query.filter(Ingreso.monto <= monto_max)
        if metodo_pago:
            ingresos_query = ingresos_query.filter(Ingreso.metodo_pago == metodo_pago)
        if buscar_texto:
            ingresos_query = ingresos_query.join(CategoriaIngreso, Ingreso.categoria_id == CategoriaIngreso.id).filter(
                or_(
                    Ingreso.notas.contains(buscar_texto),
                    CategoriaIngreso.nombre.contains(buscar_texto)
                )
            )

        resultados_ingresos = ingresos_query.order_by(Ingreso.fecha.desc()).limit(200).all()

    # Buscar egresos
    if tipo == '' or tipo == 'egreso':
        egresos_query = Egreso.query.join(Sucursal)

        # Aplicar filtros
        if sucursal_id:
            egresos_query = egresos_query.filter(Egreso.sucursal_id == sucursal_id)
        elif organizacion_id:
            sucursales_org = [s.id for s in Sucursal.query.filter_by(organizacion_id=organizacion_id).all()]
            if sucursales_org:
                egresos_query = egresos_query.filter(Egreso.sucursal_id.in_(sucursales_org))
        elif current_user.es_principal_organizacion():
            sucursales_org = [s.id for s in sucursales]
            if sucursales_org:
                egresos_query = egresos_query.filter(Egreso.sucursal_id.in_(sucursales_org))
        elif not current_user.es_admin_despacho():
            egresos_query = egresos_query.filter(Egreso.sucursal_id == current_user.sucursal_id)

        if categoria_id:
            egresos_query = egresos_query.filter(Egreso.categoria_id == categoria_id)
        if fecha_inicio:
            egresos_query = egresos_query.filter(Egreso.fecha >= datetime.strptime(fecha_inicio, '%Y-%m-%d').date())
        if fecha_fin:
            egresos_query = egresos_query.filter(Egreso.fecha <= datetime.strptime(fecha_fin, '%Y-%m-%d').date())
        if monto_min:
            egresos_query = egresos_query.filter(Egreso.monto >= monto_min)
        if monto_max:
            egresos_query = egresos_query.filter(Egreso.monto <= monto_max)
        if metodo_pago:
            egresos_query = egresos_query.filter(Egreso.metodo_pago == metodo_pago)
        if buscar_texto:
            egresos_query = egresos_query.join(CategoriaEgreso, Egreso.categoria_id == CategoriaEgreso.id).filter(
                or_(
                    Egreso.notas.contains(buscar_texto),
                    CategoriaEgreso.nombre.contains(buscar_texto)
                )
            )

        resultados_egresos = egresos_query.order_by(Egreso.fecha.desc()).limit(200).all()

    # Calcular totales
    sum_ingresos = sum([i.monto for i in resultados_ingresos]) if resultados_ingresos else Decimal('0.00')
    sum_egresos = sum([e.monto for e in resultados_egresos]) if resultados_egresos else Decimal('0.00')

    # Agregar atributo 'tipo' a cada movimiento para el template
    for ing in resultados_ingresos:
        ing.tipo = 'ingreso'
    for egr in resultados_egresos:
        egr.tipo = 'egreso'

    # Combinar y ordenar por fecha
    movimientos_combinados = sorted(
        resultados_ingresos + resultados_egresos,
        key=lambda x: x.fecha,
        reverse=True
    )

    return render_template('movimientos/listado.html',
                         ingresos=resultados_ingresos,
                         egresos=resultados_egresos,
                         movimientos_combinados=movimientos_combinados,
                         sum_ingresos=sum_ingresos,
                         sum_egresos=sum_egresos,
                         total_ingresos=len(resultados_ingresos),
                         total_egresos=len(resultados_egresos),
                         organizaciones=organizaciones,
                         sucursales=sucursales,
                         todas_las_categorias_ingreso=todas_las_categorias_ingreso,
                         todas_las_categorias_egreso=todas_las_categorias_egreso)


@bp.route('/exportar/excel')
@login_required
def exportar_excel():
    """
    Exportar movimientos a Excel
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Obtener parámetros
    sucursal_id = request.args.get('sucursal_id', type=int)
    organizacion_id = request.args.get('organizacion_id', type=int)
    mes = request.args.get('mes', type=int)
    año = request.args.get('año', type=int, default=datetime.now().year)
    tipo = request.args.get('tipo', '')  # 'ingreso', 'egreso', o '' para ambos

    # Determinar alcance: organización o sucursal
    es_por_organizacion = False
    nombre_entidad = ""
    sucursales_ids = []

    if organizacion_id:
        # Exportar por organización (todas las sucursales)
        es_por_organizacion = True
        organizacion = Organizacion.query.get_or_404(organizacion_id)

        # Verificar permisos
        if current_user.es_principal_organizacion():
            if organizacion_id != current_user.organizacion_id:
                flash('No puedes exportar movimientos de otra organización.', 'danger')
                return redirect(url_for('reportes.index'))
        elif not current_user.es_admin_despacho():
            flash('No tienes permisos para exportar reportes de organización.', 'danger')
            return redirect(url_for('reportes.index'))

        # Obtener todas las sucursales de la organización
        sucursales = Sucursal.query.filter_by(organizacion_id=organizacion_id, activa=True).all()
        sucursales_ids = [s.id for s in sucursales]
        nombre_entidad = organizacion.nombre

        if not sucursales_ids:
            flash('La organización no tiene sucursales activas.', 'warning')
            return redirect(url_for('reportes.index'))

    elif sucursal_id:
        # Exportar por sucursal individual
        # Verificar permisos
        if current_user.es_principal_organizacion():
            sucursal = Sucursal.query.get(sucursal_id)
            if not sucursal or sucursal.organizacion_id != current_user.organizacion_id:
                flash('No puedes exportar movimientos de otra organización.', 'danger')
                return redirect(url_for('movimientos.listado'))
        elif not current_user.es_admin_despacho():
            if sucursal_id != current_user.sucursal_id:
                flash('No puedes exportar movimientos de otra sucursal.', 'danger')
                return redirect(url_for('movimientos.listado'))

        sucursal = Sucursal.query.get_or_404(sucursal_id)
        sucursales_ids = [sucursal_id]
        nombre_entidad = sucursal.nombre

    else:
        flash('Debe seleccionar una organización o sucursal.', 'warning')
        return redirect(url_for('reportes.index'))

    # Crear workbook
    wb = Workbook()

    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Hoja de Ingresos
    if tipo == '' or tipo == 'ingreso':
        ws_ingresos = wb.active
        ws_ingresos.title = "Ingresos"

        # Encabezados (agregar columna de Sucursal si es reporte de organización)
        if es_por_organizacion:
            headers_ing = ['Fecha', 'Sucursal', 'Categoría', 'Monto', 'Método de Pago', 'Usuario', 'Notas']
        else:
            headers_ing = ['Fecha', 'Categoría', 'Monto', 'Método de Pago', 'Usuario', 'Notas']

        for col, header in enumerate(headers_ing, 1):
            cell = ws_ingresos.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        # Query ingresos (filtrar por lista de sucursales)
        ingresos_query = Ingreso.query.filter(Ingreso.sucursal_id.in_(sucursales_ids))
        if mes:
            ingresos_query = ingresos_query.filter(extract('month', Ingreso.fecha) == mes)
        if año:
            ingresos_query = ingresos_query.filter(extract('year', Ingreso.fecha) == año)

        ingresos = ingresos_query.order_by(Ingreso.fecha.desc()).all()

        # Datos
        for row, ingreso in enumerate(ingresos, 2):
            col = 1
            ws_ingresos.cell(row=row, column=col, value=ingreso.fecha.strftime('%d/%m/%Y')).border = border
            col += 1

            # Agregar columna de sucursal si es reporte de organización
            if es_por_organizacion:
                ws_ingresos.cell(row=row, column=col, value=ingreso.sucursal.nombre).border = border
                col += 1

            ws_ingresos.cell(row=row, column=col, value=ingreso.categoria.nombre).border = border
            col += 1
            ws_ingresos.cell(row=row, column=col, value=float(ingreso.monto)).border = border
            ws_ingresos.cell(row=row, column=col).number_format = '$#,##0.00'
            col += 1
            ws_ingresos.cell(row=row, column=col, value=ingreso.metodo_pago.title()).border = border
            col += 1
            ws_ingresos.cell(row=row, column=col, value=ingreso.usuario_registro.nombre_completo).border = border
            col += 1
            ws_ingresos.cell(row=row, column=col, value=ingreso.notas or '').border = border

        # Total
        if ingresos:
            total_row = len(ingresos) + 2
            total_col = 3 if es_por_organizacion else 2
            monto_col = 4 if es_por_organizacion else 3
            ws_ingresos.cell(row=total_row, column=total_col, value='TOTAL:').font = Font(bold=True)
            ws_ingresos.cell(row=total_row, column=monto_col, value=sum([float(i.monto) for i in ingresos])).font = Font(bold=True)
            ws_ingresos.cell(row=total_row, column=monto_col).number_format = '$#,##0.00'

        # Ajustar anchos
        ws_ingresos.column_dimensions['A'].width = 12
        if es_por_organizacion:
            ws_ingresos.column_dimensions['B'].width = 25
            ws_ingresos.column_dimensions['C'].width = 25
            ws_ingresos.column_dimensions['D'].width = 15
            ws_ingresos.column_dimensions['E'].width = 18
            ws_ingresos.column_dimensions['F'].width = 25
            ws_ingresos.column_dimensions['G'].width = 40
        else:
            ws_ingresos.column_dimensions['B'].width = 25
            ws_ingresos.column_dimensions['C'].width = 15
            ws_ingresos.column_dimensions['D'].width = 18
            ws_ingresos.column_dimensions['E'].width = 25
            ws_ingresos.column_dimensions['F'].width = 40

    # Hoja de Egresos
    if tipo == '' or tipo == 'egreso':
        if tipo == 'egreso':
            ws_egresos = wb.active
            ws_egresos.title = "Egresos"
        else:
            ws_egresos = wb.create_sheet("Egresos")

        # Encabezados (agregar columna de Sucursal si es reporte de organización)
        if es_por_organizacion:
            headers_egr = ['Fecha', 'Sucursal', 'Categoría', 'Monto', 'Método de Pago', 'Usuario', 'Comprobante', 'Notas']
        else:
            headers_egr = ['Fecha', 'Categoría', 'Monto', 'Método de Pago', 'Usuario', 'Comprobante', 'Notas']

        for col, header in enumerate(headers_egr, 1):
            cell = ws_egresos.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        # Query egresos (filtrar por lista de sucursales)
        egresos_query = Egreso.query.filter(Egreso.sucursal_id.in_(sucursales_ids))
        if mes:
            egresos_query = egresos_query.filter(extract('month', Egreso.fecha) == mes)
        if año:
            egresos_query = egresos_query.filter(extract('year', Egreso.fecha) == año)

        egresos = egresos_query.order_by(Egreso.fecha.desc()).all()

        # Datos
        for row, egreso in enumerate(egresos, 2):
            col = 1
            ws_egresos.cell(row=row, column=col, value=egreso.fecha.strftime('%d/%m/%Y')).border = border
            col += 1

            # Agregar columna de sucursal si es reporte de organización
            if es_por_organizacion:
                ws_egresos.cell(row=row, column=col, value=egreso.sucursal.nombre).border = border
                col += 1

            ws_egresos.cell(row=row, column=col, value=egreso.categoria.nombre).border = border
            col += 1
            ws_egresos.cell(row=row, column=col, value=float(egreso.monto)).border = border
            ws_egresos.cell(row=row, column=col).number_format = '$#,##0.00'
            col += 1
            ws_egresos.cell(row=row, column=col, value=egreso.metodo_pago.title()).border = border
            col += 1
            ws_egresos.cell(row=row, column=col, value=egreso.usuario_registro.nombre_completo).border = border
            col += 1
            ws_egresos.cell(row=row, column=col, value='Sí' if egreso.comprobante_filename else 'No').border = border
            col += 1
            ws_egresos.cell(row=row, column=col, value=egreso.notas or '').border = border

        # Total
        if egresos:
            total_row = len(egresos) + 2
            total_col = 3 if es_por_organizacion else 2
            monto_col = 4 if es_por_organizacion else 3
            ws_egresos.cell(row=total_row, column=total_col, value='TOTAL:').font = Font(bold=True)
            ws_egresos.cell(row=total_row, column=monto_col, value=sum([float(e.monto) for e in egresos])).font = Font(bold=True)
            ws_egresos.cell(row=total_row, column=monto_col).number_format = '$#,##0.00'

        # Ajustar anchos
        ws_egresos.column_dimensions['A'].width = 12
        if es_por_organizacion:
            ws_egresos.column_dimensions['B'].width = 25
            ws_egresos.column_dimensions['C'].width = 25
            ws_egresos.column_dimensions['D'].width = 15
            ws_egresos.column_dimensions['E'].width = 18
            ws_egresos.column_dimensions['F'].width = 25
            ws_egresos.column_dimensions['G'].width = 15
            ws_egresos.column_dimensions['H'].width = 40
        else:
            ws_egresos.column_dimensions['B'].width = 25
            ws_egresos.column_dimensions['C'].width = 15
            ws_egresos.column_dimensions['D'].width = 18
            ws_egresos.column_dimensions['E'].width = 25
            ws_egresos.column_dimensions['F'].width = 15
            ws_egresos.column_dimensions['G'].width = 40

    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Nombre del archivo
    mes_nombre = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                  'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'][mes-1] if mes else 'Todos'
    prefix = "Consolidado_" if es_por_organizacion else ""
    filename = f"{prefix}Movimientos_{nombre_entidad}_{mes_nombre}_{año}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
